import subprocess
import json
import platform
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def buscar_erros_hoje():
    ps_command = (
        f"Get-WinEvent -FilterHashTable @{{LogName='System', 'Application'; "
        f"Level=1, 2; StartTime=(Get-Date).Date}} "
        f"-ErrorAction SilentlyContinue | "
        f"Select-Object Id, ProviderName, LevelDisplayName, "
        f"@{{Name='Data';Expression={{$_.TimeCreated.ToString('dd/MM/yyyy HH:mm:ss')}}}}, "
        f"Message | ConvertTo-Json -Compress"
    )
    
    print("[*] Coletando erros criticos e avisos do DIA DE HOJE...")
    
    try:
        result = subprocess.run(["powershell", "-Command", ps_command], capture_output=True, text=True, check=True, encoding='cp850')
        saida_json = result.stdout.strip()
        
        if not saida_json:
            print("[✓] Nenhum erro registrado no dia de hoje!")
            gerar_planilha([])
            return

        eventos = json.loads(saida_json)
        if isinstance(eventos, dict):
            eventos = [eventos]
            
        gerar_planilha(eventos)
        print(f"[!] Sucesso! Planilha gerada com {len(eventos)} erro(s).")
            
    except subprocess.CalledProcessError:
        print("[x] Falha ao executar o comando no PowerShell.")
    except json.JSONDecodeError:
        print("[x] Erro ao processar os dados. O Windows retornou um formato inesperado.")
    except Exception as e:
        print(f"[x] Erro inesperado: {e}")

def gerar_planilha(eventos):
    nome_pc = platform.node()
    data_atual_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    data_arquivo = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo = f"Relatorio_Erros_{nome_pc}_{data_arquivo}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Erros Windows"
    
    # --- INÍCIO: CABEÇALHO DA MÁQUINA ---
    ws.append(["RELATÓRIO DE DIAGNÓSTICO DO SISTEMA"])
    ws.append([f"Máquina Analisada: {nome_pc}"])
    ws.append([f"Data do Scan: {data_atual_str}"])
    ws.append([""]) # Linha em branco para separar visualmente
    
    # Mesclando células para o título e informações ficarem alinhadas
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws.merge_cells('A3:E3')
    
    # Formatando o bloco de informações superiores
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A2'].font = Font(size=11, bold=True, color="333333")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['A3'].font = Font(size=11, bold=True, color="333333")
    ws['A3'].alignment = Alignment(horizontal="left", vertical="center")
    # --- FIM: CABEÇALHO DA MÁQUINA ---

    # --- INÍCIO: TABELA DE DADOS ---
    # Cabeçalhos da tabela agora entram na linha 5
    headers = ["Data/Hora", "Nível", "ID", "Fonte", "Detalhes do Erro"]
    ws.append(headers)
    
    if not eventos:
        ws.append(["-", "-", "-", "-", "Nenhum erro registrado hoje. Sistema saudável."])
    else:
        for ev in reversed(eventos):
            ev_id = str(ev.get('Id'))
            fonte = str(ev.get('ProviderName'))
            nivel = str(ev.get('LevelDisplayName', 'Erro'))
            data_hora = str(ev.get('Data'))
            
            msg_bruta = str(ev.get('Message', 'Sem descrição.'))
            msg_limpa = " ".join(msg_bruta.split())
            
            ws.append([data_hora, nivel, ev_id, fonte, msg_limpa])

    # Estilos da Tabela
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Formatando a linha 5 (Cabeçalho da tabela com os Filtros)
    for col in range(1, 6):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Formatando as linhas de dados (A partir da linha 6)
    for row in range(6, ws.max_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in [1, 2, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            
            if row % 2 == 0:
                cell.fill = zebra_fill

    # Largura das Colunas
    widths = [20, 12, 10, 25, 90]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Adicionando Filtros na linha 5 e Congelando a partir da linha 6
    ws.auto_filter.ref = f"A5:E{ws.max_row}"
    ws.freeze_panes = "A6"
    
    wb.save(nome_arquivo)

if __name__ == "__main__":
    buscar_erros_hoje()
